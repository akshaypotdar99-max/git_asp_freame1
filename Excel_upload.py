import time

from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import os
import openpyxl

driver = webdriver.Chrome()
driver.get("https://basetenant.stg.adzu.io/#page/TEST-LICENSE-ASSETS")
driver.maximize_window()
driver.implicitly_wait(15)

driver.find_element(By.ID,"us1").send_keys("akshaypotdar99@gmail.com")
driver.find_element(By.ID,"pa1").send_keys("Bharat@123")
driver.find_element(By.CSS_SELECTOR,".btn.btn-primary.login").click()
#time.sleep(7)
#Applying explicit wait
wait = WebDriverWait(driver,10)
wait.until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"div.checkbox")))

assert driver.find_element(By.CSS_SELECTOR,"div.checkbox").is_displayed()
#Edit and Save Spreadsheet
book = openpyxl.load_workbook(r"C:\Users\akshay.potdar\Downloads\TEST_ASP_SHEET.xlsx")
sheet1 = book.active
#cell = sheet1.cell(row=2,column=1)
#print(cell.value)

asset_name_1 = sheet1.cell(row=5,column=3)
asset_name_2 = sheet1.cell(row=6,column=3)
print(asset_name_1.value)
print(asset_name_2.value)
asset_name_1.value = "test_asp_Asset_3"
asset_name_2.value = "test_asp_Asset_4"
print("after edit",asset_name_1.value)
print("after edit",asset_name_2.value)

book.save(r"C:\Users\akshay.potdar\Downloads\TEST_ASP_SHEET.xlsx")
time.sleep(3)

#upload spreadsheet
driver.find_element(By.CSS_SELECTOR,"input[id='import-metadata']").send_keys(r"C:\Users\akshay.potdar\Downloads\TEST_ASP_SHEET.xlsx")

wait.until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR,"div[class='modal fade adzu-plugins-common-dialog scroll-hidden standard-modal in'] h4[class='modal-title modal-title-icon']")))
driver.find_element(By.CSS_SELECTOR,"div[class='modal fade adzu-plugins-common-dialog scroll-hidden standard-modal in'] div[class='modal-footer'] button[type='button']").click()

time.sleep(5)
asset1_name = driver.find_element(By.XPATH,f"//span[text()='{asset_name_1.value}']")
wait.until(expected_conditions.visibility_of_element_located((By.XPATH,f"//span[text()='{asset_name_1.value}']")))
assert asset1_name.is_displayed()


